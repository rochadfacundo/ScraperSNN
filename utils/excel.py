import os
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# === Cargar zonas AMBA y barrios de CABA ===
import sys

def obtener_ruta_zonas_json():
    if getattr(sys, 'frozen', False):
        # Ejecutándose como .exe (PyInstaller)
        base_path = sys._MEIPASS
    else:
        # Ejecutándose como script normal
        base_path = os.path.abspath(".")
    return os.path.join(base_path, "assets", "zonas.json")

with open(obtener_ruta_zonas_json(), encoding="utf-8") as f:
    zonas = json.load(f)


mapa_zonas = []
barrios_caba = []
for zona, partidos in zonas.items():
    for partido, localidades in partidos.items():
        for localidad in localidades:
            if zona == "CABA":
                barrios_caba.append(localidad.strip().lower())
            mapa_zonas.append((zona, partido, localidad.strip().lower()))

# === Provincias válidas (fuera de AMBA = zona personalizada)
PROVINCIAS_ARGENTINA = [
    "Buenos Aires", "Ciudad Autónoma de Buenos Aires", "CABA", "Catamarca", "Chaco", "Chubut", "Córdoba", "Corrientes",
    "Entre Ríos", "Formosa", "Jujuy", "La Pampa", "La Rioja", "Mendoza", "Misiones", "Neuquén", "Río Negro", "Salta",
    "San Juan", "San Luis", "Santa Cruz", "Santa Fe", "Santiago del Estero", "Tierra del Fuego", "Tucumán"
]

def get_localidad_y_zona(domicilio: str, provincia: str) -> tuple[str, str]:
    provincia_normalizada = provincia.strip().title()

    if provincia_normalizada not in ["Buenos Aires", "Ciudad Autónoma De Buenos Aires", "Caba"]:
        return "--", provincia_normalizada

    domicilio_lower = domicilio.lower()
    for zona, partido, localidad in mapa_zonas:
        if localidad in domicilio_lower:
            zona_resultado = partido if zona == "Fuera de AMBA" else zona
            return localidad.title(), zona_resultado

    # Si no se detectó nada pero es CABA, devolver CABA como zona
    if provincia_normalizada in ["Ciudad Autónoma De Buenos Aires", "Caba"]:
        for barrio in barrios_caba:
            if barrio in domicilio_lower:
                return barrio.title(), "CABA"
        return "--", "CABA"

    return "--", "--"

def inferir_documento_si_falta(documento: str, cuit: str) -> str:
    if documento.strip() in ["", "0", "1"] and len(cuit) == 11:
        return cuit[2:-1]  # Cortar primeros 2 y el dígito verificador final
    return documento

# === Estilos globales ===
header_font = Font(name="Cascadia Code", bold=True, size=16)
data_font = Font(name="Cascadia Code", size=12)
center_align = Alignment(horizontal="center", vertical="center")
header_fill = PatternFill(start_color="94c484", end_color="94c484", fill_type="solid")
borde_suave = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
borde_grueso = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)

def exportar_a_excel(datos: dict, matricula: str):
    carpeta_destino = Path.home() / "Documents" / "Datos de matriculas"
    carpeta_destino.mkdir(parents=True, exist_ok=True)

    fecha_str = datetime.now().strftime("%d-%m-%Y")
    archivo = carpeta_destino / f"datos {fecha_str}.xlsx"

    # Normalizar documento
    datos["Documento"] = inferir_documento_si_falta(datos.get("Documento", ""), datos.get("CUIT", ""))

    # Insertar matrícula como primera clave
    datos_completos = {"Matrícula": matricula, **datos}

    # Agregar Localidad y Zona
    domicilio = datos_completos.get("Domicilio", "")
    provincia = datos_completos.get("Provincia", "")
    localidad, zona = get_localidad_y_zona(domicilio, provincia)
    datos_completos["Localidad"] = localidad
    datos_completos["Zona"] = zona

    # Evitar guardar si todos los datos (excepto matrícula) son "--"
    if all(v == "--" for k, v in datos_completos.items() if k != "Matrícula"):
        return

    # Eliminar campo "Domicilio"
    datos_completos.pop("Domicilio", None)

    if archivo.exists():
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultado"

        for col_index, key in enumerate(datos_completos.keys(), 1):
            cell = ws.cell(row=1, column=col_index, value=key)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = borde_grueso
        ws.row_dimensions[1].height = 30

        for col_index, (key, value) in enumerate(datos_completos.items(), 1):
            max_length = max(len(str(key)), len(str(value)))
            ws.column_dimensions[get_column_letter(col_index)].width = max_length * 1.4 + 4

    next_row = ws.max_row + 1

    for col_index, value in enumerate(datos_completos.values(), 1):
        cell = ws.cell(row=next_row, column=col_index, value=value)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = borde_suave

        col_letter = get_column_letter(col_index)
        actual_width = ws.column_dimensions[col_letter].width or 0
        nuevo_ancho = len(str(value)) * 1.4 + 4
        if nuevo_ancho > actual_width:
            ws.column_dimensions[col_letter].width = nuevo_ancho

    ws.row_dimensions[next_row].height = 30
    wb.save(archivo)
