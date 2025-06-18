import os
import re
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

borde_suave = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

borde_grueso = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)


def get_dato(soup, label):
    el = soup.find('span', string=lambda s: s and label in s)
    if el and el.parent:
        text = el.parent.get_text(separator=' ', strip=True)
        cleaned = re.sub(r'\s+', ' ', text).replace(f"{label}:", "").strip()
        if label in ["Documento", "CUIT"]:
            return re.sub(r"[^\d]", "", cleaned)
        return cleaned
    return "--"

def get_dato_fuera_de_span(soup, label):
    p_tags = soup.find_all('p', class_='margen_inf')
    for p in p_tags:
        span = p.find('span', class_='destacadopdtores')
        if span and label in span.get_text(strip=True):
            # Extraer el texto del <p> sin el contenido del <span>
            full_text = p.get_text(strip=True)
            return full_text.replace(span.get_text(strip=True), "").strip()
    return "--"


def buscar_datos(matricula):
    try:
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')

        chromedriver_path = os.path.join(os.getcwd(), "chromedriver.exe")
        service = Service(executable_path=chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)

        driver.get("https://www.ssn.gob.ar/STORAGE/REGISTROS/PRODUCTORES/PRODUCTORESACTIVOSFILTRO.ASP")

        driver.find_element(By.ID, "matricula").send_keys(matricula)
        driver.find_element(By.NAME, "Submit").click()

        time.sleep(2)
        driver.switch_to.window(driver.window_handles[1])
        html = driver.page_source
        driver.quit()

        soup = BeautifulSoup(html, "html.parser")

        return {
            "Nombre": get_dato(soup, "Nombre"),
            "Documento": get_dato(soup, "Documento"),
            "CUIT": get_dato(soup, "CUIT"),
            "Domicilio": get_dato(soup, "Domicilio"),
            "Provincia": get_dato_fuera_de_span(soup, "Provincia"),
            "Teléfonos": get_dato(soup, "Teléfonos"),
            "Email": get_dato(soup, "E-mail"),
            "Ramo": get_dato(soup, "Ramo"),
            "Nro. de Resolución": get_dato_fuera_de_span(soup, "Nro. de Resolución"),
            "Fº de Resolución": get_dato_fuera_de_span(soup, "Fº de Resolución")
        }

    except Exception as e:
        return {"Error": f"Ocurrió un error: {str(e)}"}


def exportar_a_excel(datos: dict):
    archivo = "resultado.xlsx"

    # Estilos
    header_font = Font(name="Cascadia Code", bold=True, size=16)
    data_font = Font(name="Cascadia Code", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="94c484", end_color="94c484", fill_type="solid")
    borde_suave = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    borde_mas_intenso = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultado"
        for col_index, key in enumerate(datos.keys(), 1):
            cell = ws.cell(row=1, column=col_index, value=key)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = borde_mas_intenso
        ws.row_dimensions[1].height = 30
        for col_index, (key, value) in enumerate(zip(datos.keys(), datos.values()), 1):
            max_length = max(len(str(key)), len(str(value)))
            adjusted_width = max_length * 1.4 + 4
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = adjusted_width

    next_row = ws.max_row + 1

    for col_index, value in enumerate(datos.values(), 1):
        cell = ws.cell(row=next_row, column=col_index, value=value)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = borde_suave
    ws.row_dimensions[next_row].height = 30

    wb.save(archivo)
